import asyncio
import json
import random
from asyncio import Semaphore
from decimal import Decimal

import aiohttp
import openpyxl
from eth_account import Account
from eth_account.messages import encode_defunct
from fake_useragent import UserAgent
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from web3 import AsyncWeb3, AsyncHTTPProvider


url = 'https://www.bitlayer.org/me?_data=routes%2F%28%24lang%29._app%2B%2Fme%2B%2F_index'
wBTC_contract_address = "0xfF204e2681A6fA0e2C3FaDe68a1B28fb90E4Fc5F"
wBTC_abi = [{"constant": True, "inputs": [{"name": "_owner", "type": "address"}],
             "name": "balanceOf", "outputs": [{"name": "balance", "type": "uint256"}], "payable": False,
             "stateMutability": "view", "type": "function"}]


async def make_get_request(session, url, headers=None, proxy=None, params=None):
    async with session.get(url, headers=headers, proxy=proxy, params=params) as response:
        response.raise_for_status()
        return await response.json()


async def request_post(session, url, body=None, headers=None):
    async with await session.post(url, headers=headers, data=body) as response:
        response.raise_for_status()
        return response


async def signature_key(text: str, private_KEY, web3):
    signature = web3.eth.account.sign_message(encode_defunct(AsyncWeb3.to_bytes(text=text)), f'0x{private_KEY}')
    signature_str = signature.signature.hex()
    return f'0x{signature_str}'


with open('../txt_files/keys.txt', 'r') as f:
    privates = [line.strip() for line in f]
with open('../txt_files/proxy.txt', 'r') as f:
    proxies = [line.strip() for line in f]


async def check_account(session, key, proxy):
    account = Account.from_key(key)
    address = account.address
    headers = {
        'accept': '*/*',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'no-cache',
        'dnt': '1',
        'pragma': 'no-cache',
        'priority': 'u=1, i',
        'referer': 'https://www.bitlayer.org/',
        'sec-ch-ua-mobile': '?0',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': UserAgent(os='windows').random
    }
    params = {
        '_data': 'routes/($lang)._app+/me+/_index',
    }

    async_web3 = AsyncWeb3(AsyncHTTPProvider(endpoint_uri='https://rpc.bitlayer.org', request_kwargs={"proxy": proxy}))

    signature = await signature_key('BITLAYER', key, async_web3)

    login_data = {"address": key, "signature": signature}
    data = json.dumps(login_data)

    login_response = await request_post(session, 'https://www.bitlayer.org/me/login', data, headers)
    session_cookie = login_response.headers['Set-Cookie']
    headers['Cookie'] = session_cookie

    response = await make_get_request(session, url, headers=headers, proxy=proxy, params=params)

    try:
        days_on_bitlayer = response.get('profile', {}).get('daysOnBitlayer')
        txn = response.get('profile', {}).get('txn')
        bridged_in_usd = response.get('profile', {}).get('bridgedInUsd')
        total_points = response.get('profile', {}).get('totalPoints')
        level = response.get('profile', {}).get('level')
        tasks_accomplished = response.get('profile', {}).get('tasks', {}).get('accomplished')

        completed_tasks = {}
        for task_type in ['dailyTasks', 'newRacerTasks', 'advanceTasks']:
            for task in response.get('tasks', {}).get(task_type, []):
                title = task.get('title')
                is_completed = task.get('isCompleted')
                completed_tasks[title] = is_completed

        wBTC_contract = async_web3.eth.contract(address=wBTC_contract_address, abi=wBTC_abi)
        wallet_balance_wbtc = await wBTC_contract.functions.balanceOf(address).call()
        wallet_balance_btc = await async_web3.eth.get_balance(address)

        wbtc_balance = async_web3.from_wei(wallet_balance_wbtc, 'ether')
        btc_balance = async_web3.from_wei(wallet_balance_btc, 'ether')



        return {
            'key': key,
            'address': address,
            'daysOnBitlayer': days_on_bitlayer,
            'txn': txn,
            'bridgedInUsd': bridged_in_usd,
            'totalPoints': total_points,
            'level': level,
            'tasksAccomplished': tasks_accomplished,
            'completedTasks': completed_tasks,
            'wbtc_balance': wbtc_balance,
            'btc_balance': btc_balance,
            'Balance': (wbtc_balance + btc_balance) * Decimal(59286)
        }
    except Exception as e:
        print(f"Ошибка при запросе к {url}: {e}")
        return {
            'key': key,
            'error': str(e)
        }

async def get_working_proxy(proxies):
    while True:
        rand_proxy = random.choice(proxies)

        web3 = AsyncWeb3(AsyncHTTPProvider(endpoint_uri='https://rpc.bitlayer.org', request_kwargs={"proxy": rand_proxy}))

        if await web3.is_connected():
            return rand_proxy


async def main():
    all_results = []
    semaphore = Semaphore(1000)

    async def process_account_with_semaphore(session, private, proxy, i):
        async with semaphore:
            print(i + 1)
            return await check_account(session, private, proxy)

    async with aiohttp.ClientSession() as session:
        proxy = await get_working_proxy(proxies)
        tasks = [process_account_with_semaphore(session, private, proxy, i) for i, private in enumerate(privates)]
        all_results = await asyncio.gather(*tasks)

    wb = openpyxl.Workbook()
    sheet = wb.active
    print('10%')
    headers = ['Key', 'Адрес', 'Дней на Bitlayer', 'Txn', 'Bridged in USD', 'Total Points', 'Уровень',
               'Задач выполнено', 'WBTC Balance', 'BTC Balance', 'Balance $']
    if all_results and 'completedTasks' in all_results[0]:
        headers.extend(all_results[0]['completedTasks'].keys())
    sheet.append(headers)
    print('20%')
    for i, header in enumerate(headers):
        column_letter = get_column_letter(i + 1)
        if header in ('Total Points'):
            sheet.column_dimensions[column_letter].width = 11
        elif header in ('Дней на Bitlayer', 'Txn', 'Bridged in USD', 'Уровень', 'WBTC Balance', 'BTC Balance', 'Задач выполнено'):
            sheet.column_dimensions[column_letter].width = 10
        else:
            sheet.column_dimensions[column_letter].width = 20
    print('50%')

    for row_index, row in enumerate(all_results):
        values = [
            row.get('key'),
            row.get('address'),
            row.get('daysOnBitlayer'),
            row.get('txn'),
            row.get('bridgedInUsd'),
            row.get('totalPoints'),
            row.get('level'),
            row.get('tasksAccomplished'),
            row.get('wbtc_balance'),
            row.get('btc_balance'),
            row.get('Balance')
        ]
        if 'completedTasks' in row:
            values.extend(row['completedTasks'].values())
        sheet.append(values)

        if 'completedTasks' in row:
            for col_index, task_completed in enumerate(row['completedTasks'].values()):
                cell = sheet.cell(row=row_index + 2, column=len(values) - len(
                    row['completedTasks']) + col_index + 1)
                if task_completed:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        btc_balance_column = headers.index('BTC Balance') + 1

        if 'btc_balance' in row and row['btc_balance'] < 0.000002:
            cell = sheet.cell(row=row_index + 2, column=btc_balance_column)
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


    wb.save("bitlayer_results.xlsx")
    print('Готово 100%')


if __name__ == '__main__':
    asyncio.run(main())